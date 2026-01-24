"""
Backtest Report Generator
สร้าง Report แบบสวยงาม พร้อม Charts
"""
import logging
from datetime import datetime
from typing import Optional, Dict, Any
from pathlib import Path
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)


class BacktestReporter:
    """
    Generate beautiful backtest reports
    - HTML reports with charts
    - PDF export
    - Excel export
    """
    
    def __init__(self, output_dir: str = "data/backtest_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_html_report(
        self,
        result,  # BacktestResult
        filename: Optional[str] = None
    ) -> str:
        """Generate HTML report with interactive charts"""
        from backtesting.backtest_engine import BacktestResult
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_report_{result.config.symbol}_{timestamp}.html"
        
        filepath = self.output_dir / filename
        
        # Prepare data for charts
        equity_data = self._prepare_equity_chart_data(result)
        trade_data = self._prepare_trade_data(result)
        monthly_data = self._prepare_monthly_data(result)
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Trademify Backtest Report - {result.config.symbol}</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        body {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); }}
        .glass {{ background: rgba(255, 255, 255, 0.05); backdrop-filter: blur(10px); }}
        .profit {{ color: #10b981; }}
        .loss {{ color: #ef4444; }}
    </style>
</head>
<body class="text-white min-h-screen p-8">
    <div class="max-w-7xl mx-auto">
        <!-- Header -->
        <div class="text-center mb-12">
            <h1 class="text-4xl font-bold mb-2">🧪 Trademify Backtest Report</h1>
            <p class="text-gray-400">AI-Powered Trading System Analysis</p>
        </div>
        
        <!-- Summary Cards -->
        <div class="grid grid-cols-1 md:grid-cols-4 gap-6 mb-8">
            <div class="glass rounded-xl p-6">
                <p class="text-gray-400 text-sm">Total Return</p>
                <p class="text-3xl font-bold {'profit' if result.total_return > 0 else 'loss'}">{result.total_return:+.2f}%</p>
            </div>
            <div class="glass rounded-xl p-6">
                <p class="text-gray-400 text-sm">Win Rate</p>
                <p class="text-3xl font-bold text-blue-400">{result.win_rate:.1f}%</p>
            </div>
            <div class="glass rounded-xl p-6">
                <p class="text-gray-400 text-sm">Profit Factor</p>
                <p class="text-3xl font-bold text-purple-400">{result.profit_factor:.2f}</p>
            </div>
            <div class="glass rounded-xl p-6">
                <p class="text-gray-400 text-sm">Max Drawdown</p>
                <p class="text-3xl font-bold text-red-400">{result.max_drawdown:.2f}%</p>
            </div>
        </div>
        
        <!-- Config Info -->
        <div class="glass rounded-xl p-6 mb-8">
            <h2 class="text-xl font-bold mb-4">📋 Backtest Configuration</h2>
            <div class="grid grid-cols-2 md:grid-cols-4 gap-4 text-sm">
                <div>
                    <p class="text-gray-400">Symbol</p>
                    <p class="font-bold">{result.config.symbol}</p>
                </div>
                <div>
                    <p class="text-gray-400">Timeframe</p>
                    <p class="font-bold">{result.config.timeframe}</p>
                </div>
                <div>
                    <p class="text-gray-400">Period</p>
                    <p class="font-bold">{result.config.years} years</p>
                </div>
                <div>
                    <p class="text-gray-400">Initial Balance</p>
                    <p class="font-bold">${result.config.initial_balance:,.2f}</p>
                </div>
                <div>
                    <p class="text-gray-400">Min Quality</p>
                    <p class="font-bold">{result.config.min_quality}</p>
                </div>
                <div>
                    <p class="text-gray-400">Min Pass Rate</p>
                    <p class="font-bold">{result.config.min_layer_pass_rate:.0%}</p>
                </div>
                <div>
                    <p class="text-gray-400">Max Risk/Trade</p>
                    <p class="font-bold">{result.config.max_risk_per_trade}%</p>
                </div>
                <div>
                    <p class="text-gray-400">Processing Time</p>
                    <p class="font-bold">{result.processing_time:.1f}s</p>
                </div>
            </div>
        </div>
        
        <!-- Equity Chart -->
        <div class="glass rounded-xl p-6 mb-8">
            <h2 class="text-xl font-bold mb-4">📈 Equity Curve</h2>
            <canvas id="equityChart" height="300"></canvas>
        </div>
        
        <!-- Statistics Grid -->
        <div class="grid grid-cols-1 md:grid-cols-2 gap-8 mb-8">
            <!-- Financial Results -->
            <div class="glass rounded-xl p-6">
                <h2 class="text-xl font-bold mb-4">💰 Financial Results</h2>
                <table class="w-full text-sm">
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Initial Balance</td>
                        <td class="py-2 text-right font-bold">${result.config.initial_balance:,.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Final Balance</td>
                        <td class="py-2 text-right font-bold">${result.config.initial_balance + result.total_pnl:,.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Total P&L</td>
                        <td class="py-2 text-right font-bold {'profit' if result.total_pnl > 0 else 'loss'}">${result.total_pnl:+,.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Total Return</td>
                        <td class="py-2 text-right font-bold {'profit' if result.total_return > 0 else 'loss'}">{result.total_return:+.2f}%</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Annualized Return</td>
                        <td class="py-2 text-right font-bold">{result.annualized_return:.2f}%</td>
                    </tr>
                    <tr>
                        <td class="py-2 text-gray-400">Gross Profit / Loss</td>
                        <td class="py-2 text-right font-bold">${result.gross_profit:,.2f} / ${result.gross_loss:,.2f}</td>
                    </tr>
                </table>
            </div>
            
            <!-- Trading Statistics -->
            <div class="glass rounded-xl p-6">
                <h2 class="text-xl font-bold mb-4">📊 Trading Statistics</h2>
                <table class="w-full text-sm">
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Total Trades</td>
                        <td class="py-2 text-right font-bold">{result.total_trades}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Winning / Losing</td>
                        <td class="py-2 text-right font-bold">{result.winning_trades} / {result.losing_trades}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Win Rate</td>
                        <td class="py-2 text-right font-bold">{result.win_rate:.1f}%</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Profit Factor</td>
                        <td class="py-2 text-right font-bold">{result.profit_factor:.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Avg Win / Loss</td>
                        <td class="py-2 text-right font-bold">${result.avg_win:.2f} / ${result.avg_loss:.2f}</td>
                    </tr>
                    <tr>
                        <td class="py-2 text-gray-400">Expectancy</td>
                        <td class="py-2 text-right font-bold">${result.expectancy:.2f}</td>
                    </tr>
                </table>
            </div>
            
            <!-- Risk Metrics -->
            <div class="glass rounded-xl p-6">
                <h2 class="text-xl font-bold mb-4">⚠️ Risk Metrics</h2>
                <table class="w-full text-sm">
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Max Drawdown</td>
                        <td class="py-2 text-right font-bold text-red-400">{result.max_drawdown:.2f}%</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Sharpe Ratio</td>
                        <td class="py-2 text-right font-bold">{result.sharpe_ratio:.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Calmar Ratio</td>
                        <td class="py-2 text-right font-bold">{result.calmar_ratio:.2f}</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Max Consec. Wins</td>
                        <td class="py-2 text-right font-bold text-green-400">{result.max_consecutive_wins}</td>
                    </tr>
                    <tr>
                        <td class="py-2 text-gray-400">Max Consec. Losses</td>
                        <td class="py-2 text-right font-bold text-red-400">{result.max_consecutive_losses}</td>
                    </tr>
                </table>
            </div>
            
            <!-- Time Analysis -->
            <div class="glass rounded-xl p-6">
                <h2 class="text-xl font-bold mb-4">🕐 Time Analysis</h2>
                <table class="w-full text-sm">
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Avg Holding Time</td>
                        <td class="py-2 text-right font-bold">{result.avg_holding_time:.1f} hours</td>
                    </tr>
                    <tr class="border-b border-gray-700">
                        <td class="py-2 text-gray-400">Best Trading Hours</td>
                        <td class="py-2 text-right font-bold">{', '.join(str(h) for h in result.best_trading_hours) or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td class="py-2 text-gray-400">Best Trading Days</td>
                        <td class="py-2 text-right font-bold">{', '.join(result.best_trading_days) or 'N/A'}</td>
                    </tr>
                </table>
            </div>
        </div>
        
        <!-- Win Rate Chart -->
        <div class="glass rounded-xl p-6 mb-8">
            <h2 class="text-xl font-bold mb-4">🎯 Win/Loss Distribution</h2>
            <canvas id="winLossChart" height="200"></canvas>
        </div>
        
        <!-- Recent Trades Table -->
        <div class="glass rounded-xl p-6 mb-8">
            <h2 class="text-xl font-bold mb-4">📝 Recent Trades (Last 20)</h2>
            <div class="overflow-x-auto">
                <table class="w-full text-sm">
                    <thead>
                        <tr class="border-b border-gray-700 text-gray-400">
                            <th class="py-2 text-left">ID</th>
                            <th class="py-2 text-left">Side</th>
                            <th class="py-2 text-left">Entry Time</th>
                            <th class="py-2 text-right">Entry Price</th>
                            <th class="py-2 text-right">Exit Price</th>
                            <th class="py-2 text-right">P&L</th>
                            <th class="py-2 text-left">Status</th>
                            <th class="py-2 text-right">Quality</th>
                        </tr>
                    </thead>
                    <tbody>
                        {self._generate_trade_rows(result.trades[-20:])}
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Footer -->
        <div class="text-center text-gray-500 text-sm mt-8">
            <p>Generated by Trademify AI Trading System</p>
            <p>Report generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
    
    <script>
        // Equity Chart
        const equityCtx = document.getElementById('equityChart').getContext('2d');
        new Chart(equityCtx, {{
            type: 'line',
            data: {{
                labels: {equity_data['labels']},
                datasets: [{{
                    label: 'Equity',
                    data: {equity_data['equity']},
                    borderColor: 'rgba(59, 130, 246, 1)',
                    backgroundColor: 'rgba(59, 130, 246, 0.1)',
                    fill: true,
                    tension: 0.4
                }}, {{
                    label: 'Balance',
                    data: {equity_data['balance']},
                    borderColor: 'rgba(16, 185, 129, 1)',
                    borderDash: [5, 5],
                    fill: false,
                    tension: 0.4
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{
                        labels: {{ color: '#fff' }}
                    }}
                }},
                scales: {{
                    x: {{
                        ticks: {{ color: '#999' }},
                        grid: {{ color: 'rgba(255,255,255,0.1)' }}
                    }},
                    y: {{
                        ticks: {{ color: '#999' }},
                        grid: {{ color: 'rgba(255,255,255,0.1)' }}
                    }}
                }}
            }}
        }});
        
        // Win/Loss Chart
        const winLossCtx = document.getElementById('winLossChart').getContext('2d');
        new Chart(winLossCtx, {{
            type: 'doughnut',
            data: {{
                labels: ['Winning Trades', 'Losing Trades'],
                datasets: [{{
                    data: [{result.winning_trades}, {result.losing_trades}],
                    backgroundColor: ['rgba(16, 185, 129, 0.8)', 'rgba(239, 68, 68, 0.8)'],
                    borderWidth: 0
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{
                        position: 'bottom',
                        labels: {{ color: '#fff' }}
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>
"""
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        
        logger.info(f"📄 HTML report saved: {filepath}")
        return str(filepath)
    
    def _prepare_equity_chart_data(self, result) -> Dict:
        """Prepare data for equity chart"""
        if result.equity_curve.empty:
            return {'labels': [], 'equity': [], 'balance': []}
        
        # Sample data if too many points
        df = result.equity_curve
        if len(df) > 500:
            step = len(df) // 500
            df = df.iloc[::step]
        
        labels = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in df['datetime'].tolist()]
        
        return {
            'labels': str(labels),
            'equity': str(df['equity'].tolist()),
            'balance': str(df['balance'].tolist())
        }
    
    def _prepare_trade_data(self, result) -> Dict:
        """Prepare data for trade analysis"""
        if not result.trades:
            return {}
        
        pnls = [t.pnl for t in result.trades if t.exit_time]
        
        return {
            'pnls': pnls,
            'cumulative': np.cumsum(pnls).tolist()
        }
    
    def _prepare_monthly_data(self, result) -> Dict:
        """Prepare monthly performance data"""
        if not result.trades:
            return {}
        
        monthly = {}
        for trade in result.trades:
            if trade.exit_time:
                key = trade.exit_time.strftime('%Y-%m')
                if key not in monthly:
                    monthly[key] = 0
                monthly[key] += trade.pnl
        
        return monthly
    
    def _generate_trade_rows(self, trades) -> str:
        """Generate HTML table rows for trades"""
        rows = []
        for trade in trades:
            pnl_class = 'profit' if trade.pnl > 0 else 'loss'
            side_class = 'text-green-400' if trade.side == 'BUY' else 'text-red-400'
            
            row = f"""
            <tr class="border-b border-gray-800">
                <td class="py-2">{trade.id}</td>
                <td class="py-2 {side_class}">{trade.side}</td>
                <td class="py-2">{trade.entry_time.strftime('%Y-%m-%d %H:%M') if trade.entry_time else 'N/A'}</td>
                <td class="py-2 text-right">{trade.entry_price:.5f}</td>
                <td class="py-2 text-right">{trade.exit_price:.5f if trade.exit_price else 'Open'}</td>
                <td class="py-2 text-right {pnl_class}">${trade.pnl:+.2f}</td>
                <td class="py-2">{trade.status.value if trade.status else 'Open'}</td>
                <td class="py-2 text-right">{trade.signal_quality}</td>
            </tr>
            """
            rows.append(row)
        
        return '\n'.join(rows)
    
    def generate_excel_report(self, result, filename: Optional[str] = None) -> str:
        """Generate Excel report with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not installed: pip install openpyxl")
            return ""
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_report_{result.config.symbol}_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Trademify Backtest Report", ""],
            ["", ""],
            ["Configuration", ""],
            ["Symbol", result.config.symbol],
            ["Timeframe", result.config.timeframe],
            ["Period (Years)", result.config.years],
            ["Initial Balance", result.config.initial_balance],
            ["", ""],
            ["Financial Results", ""],
            ["Total P&L", result.total_pnl],
            ["Total Return (%)", result.total_return],
            ["Annualized Return (%)", result.annualized_return],
            ["", ""],
            ["Trading Statistics", ""],
            ["Total Trades", result.total_trades],
            ["Win Rate (%)", result.win_rate],
            ["Profit Factor", result.profit_factor],
            ["Max Drawdown (%)", result.max_drawdown],
            ["Sharpe Ratio", result.sharpe_ratio],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Trades sheet
        ws_trades = wb.create_sheet("Trades")
        headers = ["ID", "Symbol", "Side", "Entry Time", "Entry Price", "Exit Time", "Exit Price", "P&L", "Status", "Quality"]
        ws_trades.append(headers)
        
        for trade in result.trades:
            ws_trades.append([
                trade.id,
                trade.symbol,
                trade.side,
                trade.entry_time.isoformat() if trade.entry_time else "",
                trade.entry_price,
                trade.exit_time.isoformat() if trade.exit_time else "",
                trade.exit_price or "",
                trade.pnl,
                trade.status.value if trade.status else "",
                trade.signal_quality
            ])
        
        # Equity sheet
        if not result.equity_curve.empty:
            ws_equity = wb.create_sheet("Equity Curve")
            ws_equity.append(["DateTime", "Balance", "Equity"])
            for _, row in result.equity_curve.iterrows():
                ws_equity.append([
                    str(row['datetime']),
                    row['balance'],
                    row['equity']
                ])
        
        wb.save(filepath)
        logger.info(f"📊 Excel report saved: {filepath}")
        return str(filepath)


if __name__ == "__main__":
    # Test report generation
    print("Report generator module loaded")
